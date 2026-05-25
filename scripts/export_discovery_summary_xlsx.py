#!/usr/bin/env python3
"""Export discovery summary to Excel Tab 1."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "artifacts/catalog/discovery_summary_data.json"
OUT = ROOT / "artifacts/catalog/DISCOVERY_SUMMARY_TAB1.xlsx"


def main() -> int:
    payload = json.loads(DATA.read_text(encoding="utf-8"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Tab 1"

    ws.append(["REDMOND DISCOVERY MATERIALS LIST"])
    ws.append([
        f"Case: {payload['case_number']}",
        f"Superseding: {payload['superseding_indictment_date']}",
        f"Discovery due: {payload['discovery_due_date']}",
    ])
    ws.append([])

    headers = ["#", "Category", "Description", "Bates Begin", "Bates End", "Date", "Source Page", "Notes"]
    ws.append(headers)
    header_row = ws.max_row
    fill = PatternFill("solid", fgColor="1E293B")
    font = Font(bold=True, color="E8EEF5")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=header_row, column=col)
        c.fill = fill
        c.font = font

    for r in payload["rows"]:
        ws.append([
            r["sort_order"],
            r["category"],
            r["description"],
            r.get("bates_begin") or "",
            r.get("bates_end") or "",
            r.get("doc_date") or "",
            r["source_page"],
            r.get("notes") or "",
        ])

    for i, w in enumerate([5, 22, 42, 18, 18, 12, 16, 36], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{header_row + 1}"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT} ({len(payload['rows'])} rows on Tab 1)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
